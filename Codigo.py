import streamlit as st
from openpyxl import load_workbook
from io import BytesIO

EXCEL_PATH = "formulario.xlsx"  # archivo incluido en el proyecto

# ============================
#   Catálogo de Encargados
# ============================
encargados = {
    "Hanzel Grillo Espinoza": "111890339",
    "Roilan Gutiérrez Cruz": "111190040",
    "Marielos Arias Thiel": "108150865",
    "Mahalaed Trujillo Chaves": "402460858",
    "Silvia Arguedas Méndez": "108200386",
}

# ============================
#   Función para llenar Excel
# ============================
def llenar_excel(datos, activos):
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active

    # ----------------------------
    #   Campos fijos (celdas combinadas)
    # ----------------------------
    ws["H9"] = datos["fecha"]                     # H9–J9 combinadas
    ws["C12"] = datos["nombre"]                  # C12–E12 combinadas
    ws["J12"] = datos["cedula"]                  # J12–K12 combinadas
    ws["D14"] = datos["calidad"]                 # D14

    # ----------------------------
    #   Activos (máximo 6)
    # ----------------------------
    for idx, activo in enumerate(activos):
        row = 20 + idx
        ws[f"B{row}"] = activo["placa"]
        ws[f"D{row}"] = activo["descripcion"]     # D–E combinadas
        ws[f"G{row}"] = activo["marca"]
        ws[f"I{row}"] = activo["modelo"]
        ws[f"K{row}"] = activo["serie"]

    # ----------------------------
    #   Unidad custodio y encargado
    # ----------------------------
    ws["E32"] = datos["unidad_custodio"]          # E32–K32
    ws["E34"] = datos["encargado_bienes"]         # E34–K34
    ws["E36"] = datos["cedula_uc"]                # E36–K36

    # ----------------------------
    #   Campo de nombre en firma
    # ----------------------------
    ws["D45"] = datos["nombre"]                   # D45–E45

    # Guardar en memoria
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ============================
#   Interfaz en Streamlit
# ============================

st.title("Formulario de Préstamo de Bienes")

nombre = st.text_input("Nombre completo")
cedula = st.text_input("Cédula o carné")
fecha = st.date_input("Fecha del préstamo")

# ----------------------------
#   Menú desplegable para Calidad
# ----------------------------
calidad = st.selectbox(
    "Calidad del solicitante",
    ["Persona estudiante", "Personal docente", "Externo"]
)

num_activos = st.number_input("Cantidad de activos", 0, 6, 1)

# ----------------------------
#   Activos dinámicos
# ----------------------------
activos = []
for i in range(num_activos):
    st.subheader(f"Activo {i+1}")
    placa = st.text_input(f"Placa {i+1}")
    descripcion = st.text_input(f"Descripción {i+1}")
    marca = st.text_input(f"Marca {i+1}")
    modelo = st.text_input(f"Modelo {i+1}")
    serie = st.text_input(f"Serie {i+1}")

    activos.append({
        "placa": placa,
        "descripcion": descripcion,
        "marca": marca,
        "modelo": modelo,
        "serie": serie
    })

# ----------------------------
#   Unidad custodio fija
# ----------------------------
unidad_custodio = "Escuela de Ingeniería Industrial"
st.text_input("Unidad custodio", unidad_custodio, disabled=True)

# ----------------------------
#   Encargado + autollenado
# ----------------------------
encargado_bienes = st.selectbox(
    "Encargado de bienes institucionales",
    ["Seleccione un encargado"] + list(encargados.keys())
)

cedula_uc = ""
if encargado_bienes != "Seleccione un encargado":
    cedula_uc = encargados[encargado_bienes]

st.text_input("Cédula del encargado", cedula_uc, disabled=True)

# ----------------------------
#   BOTÓN: generar
# ----------------------------
if st.button("Generar archivo"):
    datos = {
        "nombre": nombre,
        "cedula": cedula,
        "fecha": fecha.strftime("%d/%m/%Y"),
        "calidad": calidad,
        "unidad_custodio": unidad_custodio,
        "encargado_bienes": encargado_bienes if encargado_bienes != "Seleccione un encargado" else "",
        "cedula_uc": cedula_uc
    }

    excel_file = llenar_excel(datos, activos)
    
    st.download_button(
        "Descargar formulario",
        data=excel_file,
        file_name="formulario_prestamo_filled.xlsx"
    )
